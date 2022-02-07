import os
import yaml
import torch
import librosa
import antropy
import numpy as np
import pandas as pd
import soundfile as sf

import processing.utilities as utils
import processing.input_rep as IR

from pathlib import Path
from scipy.stats import kurtosis
from openpyxl import load_workbook
from spleeter.separator import Separator
from madmom.features.beats import RNNBeatProcessor

from processing.source_separation import wv_run_spleeter
from models.models import DS_CNN, Down_CNN

proc = RNNBeatProcessor()


def append_df_to_excel(
    filename,
    df,
    sheet_name="Sheet1",
    startrow=None,
    truncate_sheet=False,
    **to_excel_kwargs
):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs
        )
        return

    # ignore [engine] parameter if it was passed
    if "engine" in to_excel_kwargs:
        to_excel_kwargs.pop("engine")

    writer = pd.ExcelWriter(filename, engine="openpyxl", mode="a")

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

    return


def l2_l1_ratio(x):
    """
    Ratio of l2 and l1 norms of a numpy array.
    -- x : array to be studied
    """
    return np.linalg.norm(x, ord=2) / np.linalg.norm(x, ord=1)


def gini_index(x):
    """
    Gini coefficient of a numpy array.
    -- x : array to be studied
    """
    x = np.sort(x)
    k = np.arange(1, x.shape[0] + 1)
    N = x.shape[0]

    return (np.sum((2 * k - N - 1) * x)) / (N * np.sum(x))


def shannon_entropy(x):
    """
    Shannon entropy of a numpy array.
    -- x : array to be studied
    """
    d = (np.linalg.norm(x, ord=2)) ** 2
    n = x ** 2
    c = n / d
    S = c * np.log(c ** 2)

    return -np.sum(S)


def max_acf(x):
    """
    Auto-correlate embedding for maximum a second and return max AC value between 0.25 and 1s shifts.
    -- x : array to be studied
    """
    x = x - x.mean()
    ac = librosa.autocorrelate(x, max_size=250)
    ac = ac / ac[0]
    return max(ac[15:])


def stats(embedding):
    """
    Function to calculate stats on embeddings.
    -- embedding : to be studied
    """
    l2l1 = l2_l1_ratio(embedding)
    gini = gini_index(embedding)
    kurt = kurtosis(embedding)
    shan = shannon_entropy(embedding)
    appp = antropy.app_entropy(embedding)
    samp = antropy.sample_entropy(embedding)
    acff = max_acf(embedding)

    return l2l1, gini, kurt, shan, appp, samp, acff


def few_note_samba(file_path, beat_model, status, separator, spl_model, cuda_available):
    """
    Function for processing raw audio with our beat tracker from A-Z.
    -- file_path      : wav, mp3...to be processed
    -- beat_model     : to be used
    -- status         : can be 'drums', 'ros', or other
    -- separator      : Spleeter separator object
    -- spl_model      : Spleeter model name
    -- cuda_available : GPU ok or no?
    """
    signal = utils.convert_to_xxhz(file_path, 44100)
    stems = wv_run_spleeter(signal, 44100, separator, spl_model)

    anchor = None
    for name, sig in stems.items():
        if name == "drums":
            possignal = np.zeros(sig.shape)
            possignal[:, :] = sig[:, :]

        else:
            if anchor is None:
                anchor = np.zeros(sig.shape)
                anchor[:, :] = sig[:, :]

            else:
                anchor[:, :] += sig[:, :]

    anchor = utils.convert_to_mono(anchor)
    anchor = librosa.resample(anchor, 44100, 16000)
    possignal = utils.convert_to_mono(possignal)
    possignal = librosa.resample(possignal, 44100, 16000)

    vqt1 = IR.generate_XQT(anchor, 16000, "vqt")
    vqt2 = IR.generate_XQT(possignal, 16000, "vqt")

    shape1 = vqt1.shape
    shape2 = vqt2.shape

    vqt1 = torch.from_numpy(vqt1).float().reshape(1, 1, shape1[0], shape1[1])
    vqt2 = torch.from_numpy(vqt2).float().reshape(1, 1, shape2[0], shape2[1])

    if cuda_available == True:
        vqt1 = vqt1.cuda()
        vqt2 = vqt2.cuda()

    with torch.no_grad():
        if status == "drums":
            output = beat_model.pretext.postve(vqt2)

        elif status == "ros":
            output = beat_model.pretext.anchor(vqt1)

        else:
            output = beat_model(vqt1, vqt2)

    return output.squeeze(dim=0).cpu().detach().numpy()


def vanilla_samba(file_path, beat_model, cuda_available):
    """
    Function for processing raw audio with our beat tracker from A-Z.
    -- file_path      : wav, mp3...to be processed
    -- beat_model     : to be used
    -- cuda_available : GPU ok or no?
    """
    signal = utils.convert_to_xxhz(file_path, 16000)
    signal = utils.convert_to_mono(signal)
    vqt = IR.generate_XQT(signal, 16000, "vqt")
    shape = vqt.shape

    vqt = torch.from_numpy(vqt).float().reshape(1, 1, shape[0], shape[1])

    if cuda_available == True:
        vqt = vqt.cuda()

    output = beat_model(vqt)

    return output.squeeze(dim=0).cpu().detach().numpy()


def bock_rnn(file_path):
    """
    Function for processing audio through Bock's 2011 RNN.
    -- file_path : wav, mp3...to be processed
    """
    output = proc(file_path)

    return output


def gtzan_44100():
    """
    Function for re-sampling GTZAN to 44100 and saving .wav files.
    """
    al = os.listdir("GTZAN/")

    idx = 0
    for el in al:
        if "mf" in el:
            continue
        else:
            wav_fps = os.listdir("GTZAN/" + el)

            for fp in wav_fps:
                full_fp = "GTZAN/" + el + "/" + fp

                y = utils.convert_to_xxhz(full_fp, 44100)
                y = y.reshape((y.shape[0]))

                Path("GTZAN2/" + el).mkdir(parents=True, exist_ok=True)

                sf.write("GTZAN2/" + el + "/" + fp, y, 44100)

                print("{} -- Saved {}.".format(idx, "GTZAN2/" + el + "/" + fp))

                idx += 1


def check_inf(l):
    """
    Check if any list elements are inf.
    -- l : list
    """
    for el in l:
        if el == float("+inf"):
            return True

    return False


def gtzan_stats(separator, spl_model, ymldict):
    """
    Function for iterating through the Gtzan dataset and computing measurements on embeddings.
    -- separator : Spleeter separator object
    -- spl_model : Spleeter model name
    """
    # Get experiment status
    model = ymldict.get("spl_mod")
    status = ymldict.get("meastatus")

    # Measures
    L2L1 = []
    GINI = []
    KURT = []
    SHAN = []
    APPP = []
    SAMP = []
    ACFF = []

    # Define model
    if status == "drums" or status == "ros" or status == "mix":
        cuda_available = torch.cuda.is_available()

        model = Down_CNN()

        state_dict = torch.load(
            "models/saved/shift_pret_cnn_16.pth", map_location=torch.device("cpu")
        )
        model.pretext.load_state_dict(state_dict)

        if cuda_available == True:
            model = model.cuda()

        model.eval()

    elif status == "van":
        cuda_available = torch.cuda.is_available()

        model = DS_CNN(pretext=True)

        state_dict = torch.load(
            "models/saved/cross_ballroom_vanilla.pth", map_location=torch.device("cpu")
        )
        model.load_state_dict(state_dict)

        if cuda_available == True:
            model = model.cuda()

        model.eval()

    elif status == "rand":
        cuda_available = torch.cuda.is_available()

        model = DS_CNN(pretext=True)

        if cuda_available == True:
            model = model.cuda()

        model.eval()

    al = os.listdir("GTZAN2/")

    idx = 0
    for el in al:
        if "mf" in el:
            continue
        else:
            wav_fps = os.listdir("GTZAN2/" + el)

            for fp in wav_fps:
                full_fp = "GTZAN2/" + el + "/" + fp

                if status == "drums" or status == "ros" or status == "mix":
                    out = few_note_samba(
                        full_fp, model, status, separator, spl_model, cuda_available
                    )

                elif status == "van" or status == "rand":
                    out = vanilla_samba(full_fp, model, cuda_available)

                else:
                    out = bock_rnn(full_fp)

                l2l1, gini, kurt, shan, appp, samp, acff = stats(out)

                ll = [l2l1, gini, kurt, shan, appp, samp, acff]
                inf_status = check_inf(ll)

                if inf_status == False:
                    L2L1.append(l2l1)
                    GINI.append(gini)
                    KURT.append(kurt)
                    SHAN.append(shan)
                    APPP.append(appp)
                    SAMP.append(samp)
                    ACFF.append(acff)

                print(
                    "{} -- L2L1: {:.8f}, GINI: {:.3f}, KURT: {:.3f}, SHAN: {:.3f}, APPP: {:.3f}, SAMP: {:.3f}, ACFF: {:.3f}.".format(
                        idx, l2l1, gini, kurt, shan, appp, samp, acff
                    )
                )

                idx += 1

    data = {
        "row1": [
            np.quantile(L2L1, 0.1),
            np.quantile(GINI, 0.1),
            np.quantile(KURT, 0.1),
            np.quantile(SHAN, 0.1),
            np.quantile(APPP, 0.1),
            np.quantile(SAMP, 0.1),
            np.quantile(ACFF, 0.1),
        ],
        "row2": [
            np.quantile(L2L1, 0.25),
            np.quantile(GINI, 0.25),
            np.quantile(KURT, 0.25),
            np.quantile(SHAN, 0.25),
            np.quantile(APPP, 0.25),
            np.quantile(SAMP, 0.25),
            np.quantile(ACFF, 0.25),
        ],
        "row3": [
            np.quantile(L2L1, 0.5),
            np.quantile(GINI, 0.5),
            np.quantile(KURT, 0.5),
            np.quantile(SHAN, 0.5),
            np.quantile(APPP, 0.5),
            np.quantile(SAMP, 0.5),
            np.quantile(ACFF, 0.5),
        ],
        "row4": [
            np.quantile(L2L1, 0.75),
            np.quantile(GINI, 0.75),
            np.quantile(KURT, 0.75),
            np.quantile(SHAN, 0.75),
            np.quantile(APPP, 0.75),
            np.quantile(SAMP, 0.75),
            np.quantile(ACFF, 0.75),
        ],
        "row5": [
            np.quantile(L2L1, 0.9),
            np.quantile(GINI, 0.9),
            np.quantile(KURT, 0.9),
            np.quantile(SHAN, 0.9),
            np.quantile(APPP, 0.9),
            np.quantile(SAMP, 0.9),
            np.quantile(ACFF, 0.9),
        ],
        "row6": [
            np.mean(L2L1),
            np.mean(GINI),
            np.mean(KURT),
            np.mean(SHAN),
            np.mean(APPP),
            np.mean(SAMP),
            np.mean(ACFF),
        ],
    }

    df = pd.DataFrame(data).T
    df = df.round(6)

    if status == "drums":
        append_df_to_excel(
            "results/measures.xlsx",
            df,
            sheet_name="Sheet1",
            startrow=7,
            startcol=2,
            truncate_sheet=False,
            engine="openpyxl",
            float_format="%.6f",
            header=False,
            index=False,
        )

    elif status == "ros":
        append_df_to_excel(
            "results/measures.xlsx",
            df,
            sheet_name="Sheet1",
            startrow=13,
            startcol=2,
            truncate_sheet=False,
            engine="openpyxl",
            float_format="%.6f",
            header=False,
            index=False,
        )

    elif status == "mix":
        append_df_to_excel(
            "results/measures.xlsx",
            df,
            sheet_name="Sheet1",
            startrow=19,
            startcol=2,
            truncate_sheet=False,
            engine="openpyxl",
            float_format="%.6f",
            header=False,
            index=False,
        )

    elif status == "van":
        append_df_to_excel(
            "results/measures.xlsx",
            df,
            sheet_name="Sheet1",
            startrow=25,
            startcol=2,
            truncate_sheet=False,
            engine="openpyxl",
            float_format="%.6f",
            header=False,
            index=False,
        )

    elif status == "rand":
        append_df_to_excel(
            "results/measures.xlsx",
            df,
            sheet_name="Sheet1",
            startrow=1,
            startcol=2,
            truncate_sheet=False,
            engine="openpyxl",
            float_format="%.6f",
            header=False,
            index=False,
        )

    return


if __name__ == "__main__":
    # Load YAML file configuations
    stream = open("configuration/config.yaml", "r")
    ymldict = yaml.safe_load(stream)
    save = ymldict.get("measave")

    if save == True:
        # Load the separation model:
        model = ymldict.get("spl_mod")
        m = "spleeter:{}".format(model)
        separator = Separator(m)

        # Compute stats on GTZAN
        gtzan_stats(separator, model, ymldict)

    else:
        # Save 44100 version of GTZAN dataset
        gtzan_44100()
